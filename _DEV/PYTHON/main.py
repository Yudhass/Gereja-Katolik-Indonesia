import os
import glob
import openpyxl
import importlib
import allow_data

kata_kunci = allow_data.kata_kunci

folder_data_mentah = "data_mentah"
folder_data_matang = "data_matang"
os.makedirs(folder_data_matang, exist_ok=True)

file_xlsx = sorted(glob.glob(os.path.join(folder_data_mentah, "*.xlsx")))

wb_katolik = openpyxl.Workbook()
ws_katolik = wb_katolik.active
ws_katolik.title = "Gereja-Katolik"

wb_luar = openpyxl.Workbook()
ws_luar = wb_luar.active
ws_luar.title = "Gereja-Luar-Konteks"

header_tertulis = False

for i, filepath in enumerate(file_xlsx, 1):
    filename = os.path.basename(filepath)
    print(f"\n{'='*60}")
    print(f"File {i}: {filename}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Sheet: {sheet_name}")
        print(f"  Baris: {ws.max_row}, Kolom: {ws.max_column}")
        for j, row in enumerate(ws.iter_rows(values_only=True)):
            print(f"    {row}")
            if j == 0:
                if not header_tertulis:
                    ws_katolik.append(row)
                    ws_luar.append(row)
                    header_tertulis = True
            else:
                name = str(row[1]).lower() if len(row) > 1 else ""

                if any(k in name for k in kata_kunci):
                    ws_katolik.append(row)
                else:
                    ws_luar.append(row)
    wb.close()

try:
    wb_katolik.save(os.path.join(folder_data_matang, "Gereja-Katolik.xlsx"))
except PermissionError:
    import time
    time.sleep(2)
    wb_katolik.save(os.path.join(folder_data_matang, "Gereja-Katolik.xlsx"))

try:
    wb_luar.save(os.path.join(folder_data_matang, "Gereja-Luar-Konteks.xlsx"))
except PermissionError:
    import time
    time.sleep(2)
    wb_luar.save(os.path.join(folder_data_matang, "Gereja-Luar-Konteks.xlsx"))
print(f"\nData gereja katolik -> data_matang/Gereja-Katolik.xlsx")
print(f"Data lainnya       -> data_matang/Gereja-Luar-Konteks.xlsx")